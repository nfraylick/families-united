import os
from google.cloud import storage
import openpyxl as xl 
bucket_name = 'ukrdctors-bucket-1'

PATH = os.path.join(os.getcwd(),'focused-century-346719-a143bf57bcc6.json' )
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = PATH

storage_client = storage.Client()

bucket = storage_client.bucket(bucket_name)

def downloadfile_from_bucket(blob_name,file_path,bucket_name):
    try:
        bucket = storage_client.get_bucket(bucket_name)
        blob = bucket.blob(blob_name)
        with open(file_path, 'wb') as f:
            storage_client.download_blob_to_file(blob,f)
        return True
    except Exception as e:
        print(e)
        return False
        
bucket_name = 'ukrdctors-bucket-1'
downloadfile_from_bucket('Physician1-Ale.xlsx', os.path.join(os.getcwd(), 'file1.xlsx'), bucket_name) 
downloadfile_from_bucket('Physician2-Nyk.xlsx', os.path.join(os.getcwd(), 'file2.xlsx'), bucket_name) 
downloadfile_from_bucket('Physician3-Tat.xlsx', os.path.join(os.getcwd(), 'file3.xlsx'), bucket_name) 


workbook = xl.load_workbook(filename = "file1.xlsx")

sheet = workbook['Sheet1']

print(sheet['A1'].value+ ': '+sheet['B1'].value)
print(sheet['A2'].value+ ': '+sheet['B2'].value)
print(sheet['A3'].value+ ': '+sheet['B3'].value)
print(sheet['A4'].value+ ': '+sheet['B4'].value)
print(str(sheet['A5'].value)+': '+ str(sheet['B5'].value))
print(sheet['A6'].value+': '+sheet['B6'].value)

workbook = xl.load_workbook(filename = "file2.xlsx")

sheet = workbook['Sheet1']

print(sheet['A1'].value+ ': '+sheet['B1'].value)
print(sheet['A2'].value+ ': '+sheet['B2'].value)
print(sheet['A3'].value+ ': '+sheet['B3'].value)
print(sheet['A4'].value+ ': '+sheet['B4'].value)
print(str(sheet['A5'].value)+': '+ str(sheet['B5'].value))
print(sheet['A6'].value+': '+sheet['B6'].value)

workbook = xl.load_workbook(filename = "file2.xlsx")

sheet = workbook['Sheet1']

print(sheet['A1'].value+ ': '+sheet['B1'].value)
print(sheet['A2'].value+ ': '+sheet['B2'].value)
print(sheet['A3'].value+ ': '+sheet['B3'].value)
print(sheet['A4'].value+ ': '+sheet['B4'].value)
print(str(sheet['A5'].value)+': '+ str(sheet['B5'].value))
print(sheet['A6'].value+': '+sheet['B6'].value)